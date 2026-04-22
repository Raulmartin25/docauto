"""
excel_generator.py
------------------
Takes structured data (from parser.py) and writes a formatted Excel file.
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from plan_categories import CATEGORY_COLUMNS


# One Movistar-only column per category, built from the shared CATEGORY_COLUMNS
# mapping so parser and Excel stay in lock-step.
_CAT_COLS = [
    (f"{short} S/ (sin IGV)", 18, True, key, {"Claro"})
    for _full, short, key in CATEGORY_COLUMNS
]

# (column label, width, is_auto_extracted, data_key, skip_for_carriers)
# skip_for_carriers: set of operador values that should NOT include this column
COLUMNS = [
    # True  = Auto-extracted from the PDF
    # False = Fill manually
    ("FECHA DE RECEPCIÓN",          16,  True,  "fecha_recepcion",            set()),
    ("LÍNEA",                       14,  True,  "numero_linea",               set()),
    ("FECHA DE SALIDA O INGRESO",   22, False,  "fecha_salida_ingreso",       set()),
    ("OPERADOR",                    13,  True,  "operador",                   set()),
    ("TIPO DE EQUIPO",              16, False,  "tipo_equipo",                set()),
    ("INICIO DE CONTRATO",          18, False,  "inicio_contrato",            set()),
    ("USUARIO",                     22, False,  "usuario",                    set()),
    ("CARGO",                       16, False,  "cargo_puesto",               set()),
    ("DNI",                         12, False,  "dni",                        set()),
    ("ÁREA",                        16, False,  "area",                       set()),
    ("OBRA",                        22, False,  "obra",                       set()),
    ("MARCA",                       13, False,  "marca",                      set()),
    ("PLAN",                        38,  True,  "plan",                       set()),
    # Movistar: one column per category replaces CARGO MENSUAL/DESCUENTOS/INAFECTO.
    # Claro: keeps the original flat layout below.
    *_CAT_COLS,
    ("CARGO MENSUAL S/ (sin IGV)",              22,  True,  "cargo_mensual",           {"Movistar"}),
    ("SERVICIOS ADICIONALES S/ (sin IGV)",     22,  True,  "servicios_adicionales",   {"Movistar"}),
    ("DESCUENTOS S/",                          16,  True,  "descuentos",              {"Claro", "Movistar"}),
    ("CARGO ADICIONAL INAFECTO S/ (sin IGV)",  22,  True,  "cargo_adicional_inafecto",{"Claro", "Movistar"}),
    ("TOTAL LÍNEA S/ (sin IGV)",        20,  True,  "total_linea",            set()),
]

_NUMERIC_KEYS = {
    "cargo_mensual", "servicios_adicionales", "descuentos",
    "cargo_adicional_inafecto", "total_linea",
    *[key for _, _, key in CATEGORY_COLUMNS],
}

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _center():
    return Alignment(horizontal="center", vertical="center")

def generate_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Telefonía"

    h      = data["header"]
    lineas = data["lineas"]

    # Filter out columns that are not relevant for this carrier
    operador    = h.get("operador", "")
    active_cols = [col for col in COLUMNS if operador not in col[4]]
    n_cols      = len(active_cols)
    last        = get_column_letter(n_cols)

    # ── Pre-compute style objects (reused across all cells) ───────────────
    border         = _border()
    fill_title     = _fill("D6E4F0")
    fill_hdr_auto  = _fill("C00000")
    fill_navy      = _fill("1F4E79")                      # header cols + total row
    fill_data_auto = (_fill("FCE4D6"), _fill("FADADD"))   # (even, odd)
    fill_data_man  = (_fill("EBF3FB"), _fill("FFFFFF"))   # filled manual (even, odd)
    fill_yellow    = _fill("FFFFC0")                      # empty manual cell
    center         = _center()
    vcenter        = Alignment(vertical="center")

    # ── Row 1 – Title ─────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last}1")
    ws["A1"]           = f"CONTROL DE TELEFONÍA MÓVIL  ·  {h['empresa']}"
    ws["A1"].font      = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = center
    ws["A1"].fill      = fill_title
    ws.row_dimensions[1].height = 28

    # ── Row 2 – Receipt metadata ──────────────────────────────────────────
    meta = [
        ("Operador",  h["operador"]),
        ("Recibo N°", h["n_recibo"]),
        ("Emisión",   h["fecha_emision"]),
        ("Período",   h["periodo"]),
        ("RUC",       h["ruc"]),
    ]
    col = 1
    for label, value in meta:
        if col > n_cols:
            break                    # guard: don't write past the last column
        ws.cell(2, col, label).font = Font(bold=True, size=9, color="1F4E79")
        ws.cell(2, col, label).fill = fill_title
        if col + 1 <= n_cols:
            ws.cell(2, col + 1, value).font = Font(size=9)
            ws.cell(2, col + 1, value).fill = fill_title
        col += 2
    if col <= n_cols:
        ws.merge_cells(f"{get_column_letter(col)}2:{last}2")
        ws.cell(2, col).fill = fill_title
    ws.row_dimensions[2].height = 18

    # ── Row 3 – Legend ────────────────────────────────────────────────────
    ws.merge_cells(f"A3:{last}3")
    ws["A3"] = (
        "  Columnas AMARILLAS → completar manualmente   |   "
        "Columnas ROJAS → extraídas automáticamente del recibo PDF"
    )
    ws["A3"].font      = Font(italic=True, size=8, color="595959")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 14

    # ── Row 4 – Column headers ────────────────────────────────────────────
    HDR_ROW = 4
    for ci, (label, width, is_auto, _key, _skip) in enumerate(active_cols, 1):
        cell           = ws.cell(HDR_ROW, ci, label)
        cell.fill      = fill_hdr_auto if is_auto else fill_navy
        cell.font      = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR_ROW].height = 32

    # Precompute which column indices (1-based) are numeric
    numeric_ci = {ci for ci, (_, _, _, key, _) in enumerate(active_cols, 1) if key in _NUMERIC_KEYS}

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, linea in enumerate(lineas):
        row_data = []
        for _, _, _, key, _ in active_cols:
            val = linea.get(key, "")
            if key in _NUMERIC_KEYS and val != "":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            row_data.append(val)
        ws.append(row_data)
        dr  = ws.max_row
        alt = i % 2 == 0

        for ci, (_, _, is_auto, _key, _) in enumerate(active_cols, 1):
            cell           = ws.cell(dr, ci)
            cell.border    = border
            cell.alignment = vcenter
            if is_auto:
                cell.fill = fill_data_auto[0] if alt else fill_data_auto[1]
                if ci in numeric_ci:
                    cell.number_format = "#,##0.00"
            else:
                cell.fill = fill_yellow if cell.value == "" else (fill_data_man[0] if alt else fill_data_man[1])

    # ── Totals row ────────────────────────────────────────────────────────
    ds = HDR_ROW + 1
    de = ws.max_row
    tr = ws.max_row + 1

    ws.cell(tr, 1, "TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(tr, 1).fill          = fill_navy
    ws.cell(tr, 1).alignment     = center
    ws.cell(tr, 1).border        = border
    ws.cell(tr, 2, f"{len(lineas)} líneas").font = Font(bold=True, color="FFFFFF")
    ws.cell(tr, 2).fill          = fill_navy
    ws.cell(tr, 2).alignment     = center
    ws.cell(tr, 2).border        = border

    for ci in range(3, n_cols + 1):
        cell = ws.cell(tr, ci)
        if ci in numeric_ci:
            cell.value         = f"=SUM({get_column_letter(ci)}{ds}:{get_column_letter(ci)}{de})"
            cell.number_format = "#,##0.00"
        cell.fill      = fill_navy
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border    = border

    # ── Redondeo + Total final (last column) ──────────────────────────────
    # Only carriers that expose a Redondeo line on the receipt (Movistar today)
    # need these extra rows — they reconcile the per-line sum with the
    # subtotal printed on the PDF.
    redondeo = h.get("redondeo")
    if redondeo is not None:
        total_ci = n_cols  # TOTAL LÍNEA is always the last active column
        total_letter = get_column_letter(total_ci)
        fill_redondeo = _fill("FFF2CC")
        fill_total_final = _fill("E2EFDA")

        rr = tr + 1
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=total_ci - 1)
        ws.cell(rr, 1, "REDONDEO").font      = Font(bold=True, color="1F4E79")
        ws.cell(rr, 1).alignment             = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, total_ci):
            ws.cell(rr, ci).fill   = fill_redondeo
            ws.cell(rr, ci).border = border
        amt_cell = ws.cell(rr, total_ci, float(redondeo))
        amt_cell.number_format = "#,##0.00"
        amt_cell.font          = Font(bold=True, color="1F4E79")
        amt_cell.alignment     = Alignment(horizontal="right", vertical="center")
        amt_cell.fill          = fill_redondeo
        amt_cell.border        = border

        fr = tr + 2
        ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=total_ci - 1)
        ws.cell(fr, 1, "TOTAL").font = Font(bold=True, color="375623")
        ws.cell(fr, 1).alignment     = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, total_ci):
            ws.cell(fr, ci).fill   = fill_total_final
            ws.cell(fr, ci).border = border
        final_cell = ws.cell(fr, total_ci)
        final_cell.value         = f"={total_letter}{tr}+{total_letter}{rr}"
        final_cell.number_format = "#,##0.00"
        final_cell.font          = Font(bold=True, color="375623")
        final_cell.alignment     = Alignment(horizontal="right", vertical="center")
        final_cell.fill          = fill_total_final
        final_cell.border        = border

    ws.freeze_panes = f"A{HDR_ROW + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
